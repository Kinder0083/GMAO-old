"""
Service de planification et envoi automatique des rapports M.E.S.
Utilise APScheduler pour programmer l'envoi des rapports par email
"""
import asyncio
import logging
import io
from datetime import datetime, timezone, timedelta
from typing import Optional
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from bson import ObjectId

logger = logging.getLogger(__name__)

# Global scheduler instance
mes_report_scheduler = None


class MESReportScheduler:
    """Scheduler for automatic M.E.S. report sending"""
    
    def __init__(self, db, mes_service, email_service_module):
        self.db = db
        self.mes_service = mes_service
        self.email_service = email_service_module
        self.scheduler = AsyncIOScheduler()
        self._jobs = {}  # report_id -> job_id mapping
    
    async def start(self):
        """Start the scheduler and load existing scheduled reports"""
        logger.info("[MES Report Scheduler] Demarrage...")
        
        # Load all active scheduled reports
        reports = await self.db.mes_scheduled_reports.find({"active": True}).to_list(100)
        
        for report in reports:
            self._add_job(report)
        
        self.scheduler.start()
        logger.info(f"[MES Report Scheduler] {len(reports)} rapport(s) planifie(s)")
    
    def stop(self):
        """Stop the scheduler"""
        if self.scheduler.running:
            self.scheduler.shutdown()
            logger.info("[MES Report Scheduler] Arrete")
    
    def _add_job(self, report: dict):
        """Add a scheduled job for a report"""
        report_id = str(report["_id"])
        frequency = report.get("frequency", "weekly")
        hour = report.get("hour", 8)
        minute = report.get("minute", 0)
        
        # Build cron trigger based on frequency
        if frequency == "daily":
            trigger = CronTrigger(hour=hour, minute=minute)
        elif frequency == "weekly":
            day_of_week = report.get("day_of_week", 0)  # 0=Monday
            trigger = CronTrigger(day_of_week=day_of_week, hour=hour, minute=minute)
        elif frequency == "monthly":
            day = report.get("day_of_month", 1)
            trigger = CronTrigger(day=day, hour=hour, minute=minute)
        else:
            logger.warning(f"[MES Report Scheduler] Frequence inconnue: {frequency}")
            return
        
        job = self.scheduler.add_job(
            self._send_report_job,
            trigger,
            args=[report_id],
            id=f"mes_report_{report_id}",
            replace_existing=True,
            name=f"MES Report: {report.get('name', report_id)}"
        )
        
        self._jobs[report_id] = job.id
        logger.info(f"[MES Report Scheduler] Job ajoute: {report.get('name')} ({frequency})")
    
    def _remove_job(self, report_id: str):
        """Remove a scheduled job"""
        job_id = self._jobs.get(report_id)
        if job_id:
            try:
                self.scheduler.remove_job(job_id)
                del self._jobs[report_id]
                logger.info(f"[MES Report Scheduler] Job supprime: {report_id}")
            except Exception as e:
                logger.warning(f"[MES Report Scheduler] Erreur suppression job: {e}")
    
    async def refresh_job(self, report_id: str):
        """Refresh a job (after update)"""
        self._remove_job(report_id)
        report = await self.db.mes_scheduled_reports.find_one({"_id": ObjectId(report_id)})
        if report and report.get("active"):
            self._add_job(report)
    
    async def _send_report_job(self, report_id: str):
        """Job function that sends the report"""
        try:
            await self.send_report(report_id)
        except Exception as e:
            logger.error(f"[MES Report Scheduler] Erreur envoi rapport {report_id}: {e}")
    
    async def send_report(self, report_id: str):
        """Generate and send a scheduled report"""
        report = await self.db.mes_scheduled_reports.find_one({"_id": ObjectId(report_id)})
        if not report:
            logger.warning(f"[MES Report Scheduler] Rapport non trouve: {report_id}")
            return
        
        logger.info(f"[MES Report Scheduler] Generation rapport: {report.get('name')}")
        
        # Calculate date range based on frequency
        now = datetime.now(timezone.utc)
        frequency = report.get("frequency", "weekly")
        
        if frequency == "daily":
            # Yesterday
            date_from = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = (now - timedelta(days=1)).replace(hour=23, minute=59, second=59)
        elif frequency == "weekly":
            # Last 7 days
            date_from = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = (now - timedelta(days=1)).replace(hour=23, minute=59, second=59)
        elif frequency == "monthly":
            # Last month
            first_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            date_to = first_of_month - timedelta(seconds=1)
            date_from = date_to.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            date_from = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = now
        
        # Get report data
        report_data = await self.mes_service.get_report_data(
            report.get("machine_ids", ["all"]),
            report.get("report_type", "all"),
            date_from.isoformat(),
            date_to.isoformat()
        )
        
        # Generate file
        file_format = report.get("format", "pdf")
        if file_format == "pdf":
            file_buffer = await self._generate_pdf(report_data, date_from, date_to, report.get("report_type", "all"))
            filename = f"rapport_mes_{date_from.strftime('%Y-%m-%d')}_{date_to.strftime('%Y-%m-%d')}.pdf"
            mime_type = "application/pdf"
        else:
            file_buffer = await self._generate_excel(report_data, date_from, date_to, report.get("report_type", "all"))
            filename = f"rapport_mes_{date_from.strftime('%Y-%m-%d')}_{date_to.strftime('%Y-%m-%d')}.xlsx"
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        # Send email to all recipients
        recipients = report.get("recipients", [])
        if not recipients:
            logger.warning(f"[MES Report Scheduler] Pas de destinataires pour {report.get('name')}")
            return
        
        subject = f"[FSAO] {report.get('name', 'Rapport M.E.S.')} - {date_from.strftime('%d/%m/%Y')} au {date_to.strftime('%d/%m/%Y')}"
        
        # Build HTML body
        summary = report_data.get("summary", {})
        body_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <h2 style="color: #4F46E5;">{report.get('name', 'Rapport M.E.S.')}</h2>
            <p>Bonjour,</p>
            <p>Veuillez trouver ci-joint le rapport M.E.S. automatique pour la periode du <strong>{date_from.strftime('%d/%m/%Y')}</strong> au <strong>{date_to.strftime('%d/%m/%Y')}</strong>.</p>
            
            <h3 style="color: #6366F1;">Resume</h3>
            <table style="border-collapse: collapse; margin: 10px 0;">
                <tr><td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>Machines</strong></td><td style="padding: 5px 15px; border: 1px solid #ddd;">{summary.get('total_machines', 0)}</td></tr>
                <tr><td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>Production totale</strong></td><td style="padding: 5px 15px; border: 1px solid #ddd;">{summary.get('total_production', 0)}</td></tr>
                <tr><td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>Rebuts</strong></td><td style="padding: 5px 15px; border: 1px solid #ddd;">{summary.get('total_rejects', 0)}</td></tr>
                <tr><td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>Temps d'arret</strong></td><td style="padding: 5px 15px; border: 1px solid #ddd;">{summary.get('total_downtime_hours', 0)} h</td></tr>
                <tr><td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>TRS moyen</strong></td><td style="padding: 5px 15px; border: 1px solid #ddd;">{summary.get('average_trs', 0)} %</td></tr>
            </table>
            
            <p>Consultez le fichier joint pour les details complets.</p>
            <p style="color: #888; font-size: 12px; margin-top: 30px;">
                Ce rapport a ete genere automatiquement par FSAO Iris.<br>
                Pour modifier ou desactiver ce rapport, connectez-vous a l'application.
            </p>
        </body>
        </html>
        """
        
        # Send email with attachment
        for recipient in recipients:
            try:
                await self._send_email_with_attachment(
                    recipient, subject, body_html, 
                    file_buffer.getvalue(), filename, mime_type
                )
                logger.info(f"[MES Report Scheduler] Email envoye a {recipient}")
            except Exception as e:
                logger.error(f"[MES Report Scheduler] Erreur envoi email a {recipient}: {e}")
        
        # Mark report as sent
        await self.mes_service.mark_report_sent(report_id)
        logger.info(f"[MES Report Scheduler] Rapport envoye: {report.get('name')}")
    
    async def _generate_pdf(self, report_data, date_from, date_to, report_type):
        """Generate PDF report"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=20)
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, alignment=TA_LEFT, spaceAfter=10)
        normal_style = styles['Normal']
        
        elements = []
        
        # Title
        elements.append(Paragraph("RAPPORT M.E.S. AUTOMATIQUE", title_style))
        elements.append(Paragraph(f"Periode: {date_from.strftime('%d/%m/%Y')} - {date_to.strftime('%d/%m/%Y')}", normal_style))
        elements.append(Paragraph(f"Genere le: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}", normal_style))
        elements.append(Spacer(1, 20))
        
        # Summary
        summary = report_data.get("summary", {})
        elements.append(Paragraph("RESUME GLOBAL", subtitle_style))
        summary_data = [
            ["Indicateur", "Valeur"],
            ["Nombre de machines", str(summary.get("total_machines", 0))],
            ["Production totale", str(summary.get("total_production", 0))],
            ["Rebuts totaux", str(summary.get("total_rejects", 0))],
            ["Temps d'arret total", f"{summary.get('total_downtime_hours', 0)} h"],
            ["TRS moyen", f"{summary.get('average_trs', 0)} %"],
        ]
        summary_table = Table(summary_data, colWidths=[8*cm, 5*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # TRS Table
        if report_type in ["trs", "all"]:
            elements.append(Paragraph("TRS PAR MACHINE", subtitle_style))
            trs_rows = [["Machine", "TRS Moy.", "Dispo.", "Perf.", "Qualite"]]
            for machine in report_data.get("machines", []):
                trs_data = machine.get("trs", {})
                trs_rows.append([
                    machine.get("name", "")[:30],
                    f"{trs_data.get('average_trs', 0)} %",
                    f"{trs_data.get('average_availability', 0)} %",
                    f"{trs_data.get('average_performance', 0)} %",
                    f"{trs_data.get('average_quality', 0)} %",
                ])
            if len(trs_rows) > 1:
                trs_table = Table(trs_rows, colWidths=[6*cm, 3*cm, 3*cm, 3*cm, 3*cm])
                trs_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(trs_table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    async def _generate_excel(self, report_data, date_from, date_to, report_type):
        """Generate Excel report"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Resume"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        
        # Summary
        ws.append(["RAPPORT M.E.S. AUTOMATIQUE"])
        ws.append([f"Periode: {date_from.strftime('%d/%m/%Y')} - {date_to.strftime('%d/%m/%Y')}"])
        ws.append([])
        
        summary = report_data.get("summary", {})
        ws.append(["Indicateur", "Valeur"])
        ws["A4"].font = header_font
        ws["A4"].fill = header_fill
        ws["B4"].font = header_font
        ws["B4"].fill = header_fill
        
        ws.append(["Machines", summary.get("total_machines", 0)])
        ws.append(["Production totale", summary.get("total_production", 0)])
        ws.append(["Rebuts", summary.get("total_rejects", 0)])
        ws.append(["Temps arret (h)", summary.get("total_downtime_hours", 0)])
        ws.append(["TRS moyen (%)", summary.get("average_trs", 0)])
        
        # TRS Sheet
        if report_type in ["trs", "all"]:
            ws_trs = wb.create_sheet("TRS")
            ws_trs.append(["Machine", "TRS (%)", "Disponibilite (%)", "Performance (%)", "Qualite (%)"])
            for cell in ws_trs[1]:
                cell.font = header_font
                cell.fill = header_fill
            for machine in report_data.get("machines", []):
                trs = machine.get("trs", {})
                ws_trs.append([
                    machine.get("name", ""),
                    trs.get("average_trs", 0),
                    trs.get("average_availability", 0),
                    trs.get("average_performance", 0),
                    trs.get("average_quality", 0),
                ])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    async def _send_email_with_attachment(self, recipient, subject, body_html, attachment_data, filename, mime_type):
        """Send email with attachment using the email service"""
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email import encoders
        
        # Get SMTP config from email_service
        smtp_server = self.email_service.SMTP_SERVER
        smtp_port = self.email_service.SMTP_PORT
        smtp_username = self.email_service.SMTP_USERNAME
        smtp_password = self.email_service.SMTP_PASSWORD
        sender_email = self.email_service.SMTP_SENDER_EMAIL
        use_tls = self.email_service.SMTP_USE_TLS
        
        if not smtp_server or smtp_server == 'localhost':
            logger.warning("[MES Report Scheduler] SMTP non configure, email non envoye")
            return
        
        msg = MIMEMultipart()
        msg['From'] = f"FSAO Iris <{sender_email}>"
        msg['To'] = recipient
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body_html, 'html'))
        
        # Add attachment
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
        
        # Send
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            if use_tls:
                server.starttls()
            if smtp_username and smtp_password:
                server.login(smtp_username, smtp_password)
            server.send_message(msg)


async def init_mes_report_scheduler(db, mes_service, email_service_module):
    """Initialize and start the M.E.S. report scheduler"""
    global mes_report_scheduler
    mes_report_scheduler = MESReportScheduler(db, mes_service, email_service_module)
    await mes_report_scheduler.start()
    return mes_report_scheduler
