"""
Service pour lire les fichiers Excel via SMB/CIFS
"""
import os
import io
import logging
from typing import Optional, Dict, Any, List, Union
from datetime import datetime, timezone
import pandas as pd
from openpyxl import load_workbook
from smbclient import open_file, register_session, ClientConfig
import re

logger = logging.getLogger(__name__)

# Configuration SMB globale
_smb_config = {
    "username": None,
    "password": None,
    "domain": None
}


def configure_smb(username: str = None, password: str = None, domain: str = None):
    """Configure les credentials SMB globaux"""
    global _smb_config
    _smb_config["username"] = username or os.environ.get("SMB_USERNAME")
    _smb_config["password"] = password or os.environ.get("SMB_PASSWORD")
    _smb_config["domain"] = domain or os.environ.get("SMB_DOMAIN", "")
    
    if _smb_config["username"] and _smb_config["password"]:
        # Configuration du client SMB
        ClientConfig(username=_smb_config["username"], password=_smb_config["password"])
        logger.info(f"SMB configuré pour l'utilisateur {_smb_config['username']}")


def parse_smb_path(path: str) -> Dict[str, str]:
    """
    Parse un chemin SMB et extrait le serveur, le partage et le chemin du fichier
    
    Ex: \\\\serveur\\partage\\dossier\\fichier.xlsx
    """
    # Normaliser le chemin (remplacer / par \\)
    path = path.replace("/", "\\")
    
    # Enlever les \\\\ du début si présents
    if path.startswith("\\\\"):
        path = path[2:]
    
    parts = path.split("\\")
    
    if len(parts) < 3:
        raise ValueError(f"Chemin SMB invalide: {path}")
    
    return {
        "server": parts[0],
        "share": parts[1],
        "path": "\\".join(parts[2:]) if len(parts) > 2 else ""
    }


def read_excel_from_smb(
    smb_path: str,
    sheet_name: Optional[str] = None,
    cell_reference: Optional[str] = None,
    column_name: Optional[str] = None,
    row_filter: Optional[Dict[str, Any]] = None,
    aggregation: Optional[str] = None,
    username: Optional[str] = None,
    password: Optional[str] = None
) -> Union[float, str, List, Dict, pd.DataFrame]:
    """
    Lit des données depuis un fichier Excel via SMB/CIFS
    
    Args:
        smb_path: Chemin SMB complet (ex: \\\\serveur\\partage\\fichier.xlsx)
        sheet_name: Nom de la feuille (défaut: première feuille)
        cell_reference: Référence de cellule (ex: A1, B2:D10)
        column_name: Nom de la colonne pour récupérer une valeur
        row_filter: Filtre sur les lignes (dict {colonne: valeur})
        aggregation: Fonction d'agrégation (SUM, AVG, COUNT, MIN, MAX)
        username: Identifiant SMB (optionnel si configuré globalement)
        password: Mot de passe SMB (optionnel si configuré globalement)
    
    Returns:
        Valeur, liste de valeurs, ou DataFrame selon les paramètres
    """
    try:
        # Utiliser les credentials globaux si non fournis
        smb_user = username or _smb_config.get("username")
        smb_pass = password or _smb_config.get("password")
        
        # Parser le chemin SMB
        parsed = parse_smb_path(smb_path)
        server = parsed["server"]
        share = parsed["share"]
        file_path = parsed["path"]
        
        # Construire le chemin SMB complet
        full_smb_path = f"\\\\{server}\\{share}\\{file_path}"
        
        logger.info(f"Lecture Excel depuis: {full_smb_path}")
        
        # Ouvrir le fichier via SMB
        with open_file(full_smb_path, mode="rb", username=smb_user, password=smb_pass) as f:
            content = f.read()
        
        # Charger le fichier Excel
        excel_buffer = io.BytesIO(content)
        
        # Si une référence de cellule spécifique est demandée
        if cell_reference and ":" not in cell_reference:
            # Cellule unique (ex: A1)
            wb = load_workbook(excel_buffer, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            value = ws[cell_reference].value
            wb.close()
            return value
        
        # Charger avec pandas pour les plages et les filtres
        df = pd.read_excel(excel_buffer, sheet_name=sheet_name or 0)
        
        # Si une plage de cellules est spécifiée
        if cell_reference and ":" in cell_reference:
            # Convertir la plage en indices pandas
            start_cell, end_cell = cell_reference.split(":")
            start_col, start_row = _parse_cell_reference(start_cell)
            end_col, end_row = _parse_cell_reference(end_cell)
            
            df = df.iloc[start_row:end_row+1, start_col:end_col+1]
        
        # Appliquer le filtre de lignes
        if row_filter:
            for col, val in row_filter.items():
                if col in df.columns:
                    df = df[df[col] == val]
        
        # Si une colonne spécifique est demandée
        if column_name and column_name in df.columns:
            data = df[column_name]
            
            # Appliquer l'agrégation
            if aggregation:
                return _apply_aggregation(data, aggregation)
            
            # Retourner la première valeur ou la liste
            if len(data) == 1:
                return data.iloc[0]
            return data.tolist()
        
        # Appliquer l'agrégation sur tout le DataFrame numérique
        if aggregation:
            numeric_df = df.select_dtypes(include=['number'])
            if not numeric_df.empty:
                return _apply_aggregation(numeric_df.values.flatten(), aggregation)
        
        # Retourner le DataFrame complet en dict
        return df.to_dict(orient='records')
        
    except Exception as e:
        logger.error(f"Erreur lecture Excel SMB: {e}")
        raise Exception(f"Erreur lecture fichier Excel: {str(e)}")


def _parse_cell_reference(cell: str) -> tuple:
    """Convertit une référence de cellule Excel en indices (col, row)"""
    match = re.match(r"([A-Z]+)(\d+)", cell.upper())
    if not match:
        raise ValueError(f"Référence de cellule invalide: {cell}")
    
    col_str, row_str = match.groups()
    
    # Convertir la lettre de colonne en index
    col_idx = 0
    for char in col_str:
        col_idx = col_idx * 26 + (ord(char) - ord('A'))
    
    row_idx = int(row_str) - 1  # Excel commence à 1
    
    return col_idx, row_idx


def _apply_aggregation(data, aggregation: str) -> float:
    """Applique une fonction d'agrégation sur les données"""
    import numpy as np
    
    # Convertir en array numpy en ignorant les NaN
    arr = np.array([x for x in data if pd.notna(x) and isinstance(x, (int, float))])
    
    if len(arr) == 0:
        return 0
    
    aggregation = aggregation.upper()
    
    if aggregation == "SUM":
        return float(np.sum(arr))
    elif aggregation == "AVG" or aggregation == "AVERAGE":
        return float(np.mean(arr))
    elif aggregation == "MIN":
        return float(np.min(arr))
    elif aggregation == "MAX":
        return float(np.max(arr))
    elif aggregation == "COUNT":
        return float(len(arr))
    elif aggregation == "MEDIAN":
        return float(np.median(arr))
    elif aggregation == "STD":
        return float(np.std(arr))
    else:
        raise ValueError(f"Fonction d'agrégation inconnue: {aggregation}")


def test_smb_connection(smb_path: str, username: str = None, password: str = None) -> Dict[str, Any]:
    """
    Teste la connexion SMB et retourne des informations sur le fichier
    """
    try:
        smb_user = username or _smb_config.get("username")
        smb_pass = password or _smb_config.get("password")
        
        parsed = parse_smb_path(smb_path)
        full_smb_path = f"\\\\{parsed['server']}\\{parsed['share']}\\{parsed['path']}"
        
        with open_file(full_smb_path, mode="rb", username=smb_user, password=smb_pass) as f:
            content = f.read()
            size = len(content)
        
        # Charger pour obtenir les infos
        excel_buffer = io.BytesIO(content)
        wb = load_workbook(excel_buffer, data_only=True)
        
        sheets = wb.sheetnames
        wb.close()
        
        return {
            "success": True,
            "file_size": size,
            "sheets": sheets,
            "server": parsed["server"],
            "share": parsed["share"],
            "path": parsed["path"]
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


def get_excel_preview(
    smb_path: str,
    sheet_name: str = None,
    max_rows: int = 10,
    username: str = None,
    password: str = None
) -> Dict[str, Any]:
    """
    Retourne un aperçu du fichier Excel (premières lignes et colonnes)
    """
    try:
        smb_user = username or _smb_config.get("username")
        smb_pass = password or _smb_config.get("password")
        
        parsed = parse_smb_path(smb_path)
        full_smb_path = f"\\\\{parsed['server']}\\{parsed['share']}\\{parsed['path']}"
        
        with open_file(full_smb_path, mode="rb", username=smb_user, password=smb_pass) as f:
            content = f.read()
        
        excel_buffer = io.BytesIO(content)
        df = pd.read_excel(excel_buffer, sheet_name=sheet_name or 0, nrows=max_rows)
        
        return {
            "success": True,
            "columns": df.columns.tolist(),
            "data": df.head(max_rows).to_dict(orient='records'),
            "total_columns": len(df.columns),
            "sample_rows": len(df)
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }
